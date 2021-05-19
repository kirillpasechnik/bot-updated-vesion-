from openpyxl import load_workbook


book = load_workbook('Книга1.xlsx')
sheet_1 = book['Лист1']
stickers_page = book['Стикеры']
print(book.worksheets)
for i in range(1, 6):
    print(stickers_page.cell(row=i, column=1).value)